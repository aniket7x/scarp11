"""
MERGED SCRIPT — All 12 Report Scrapers in One
==============================================
Reports covered (login once, all run sequentially):

  e-Mutation (Defect Check):
    >10 days, >15 days, >20 days

  e-Mutation (Parimarjan — Digitization):
    >35 days, >75 days, >120 days

  e-Mutation (Parimarjan — Rectification):
    >35 days, >75 days, >120 days

  e-Mutation (General):
    >35 days, >75 days, >120 days

Output per run: one Excel + one PDF-folder per report/bucket combination,
all saved to ~/Desktop.

FIXES IN THIS VERSION:
  1. View-page scraping no longer waits up to ~50s when a page has no
     action table. It now polls fast and, once the DOM stops changing
     with no valid table found, bails out after ~1.6s. The moment a
     valid action table appears, it scrapes and closes immediately
     (unchanged fast path).
  2. Pagination control rows (e.g. "1 2 3 4 / 5 6 7 ...") no longer leak
     into the scraped e-Mutation rows / PDFs.
  3. Every karmchari now gets a sheet in the output Excel for every
     report/bucket, even if zero matching cases were found.
  4. PAGINATION ELLIPSIS FIX (this version): ASP.NET GridView renders
     the "..." pager trigger as a <span> (non-clickable) in some
     versions and as an <a> in others. This version handles all cases:
       a) <a> with text "..." or "…"
       b) <span> with text "..." — finds the closest sibling/parent <a>
          with __doPostBack in onclick
       c) Direct __doPostBack injection: parses existing pager link
          onclick values to infer the next page-group's __doPostBack
          target and argument, then calls __doPostBack directly.
     This fixes the bug visible in the screenshot where pages 1–10 are
     directly clickable but page 11 onward requires clicking "..." which
     is a <span> in this portal's ASP.NET GridView pager.
"""

import os, sys, subprocess, re, asyncio, gc, urllib.request, math, textwrap, tempfile

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ── Dependencies ────────────────────────────────────────────────────────────────
def install_dependencies():
    pkgs = ["playwright", "pandas", "openpyxl", "reportlab", "Pillow"]
    try:
        import playwright, pandas, openpyxl, reportlab, PIL
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + pkgs)
    subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])

install_dependencies()

import pandas as pd
from datetime import datetime
from io import BytesIO

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle,
                                 Spacer, HRFlowable)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ── Font setup ──────────────────────────────────────────────────────────────────
_WIN = "C:/Windows/Fonts"
_FONT_CANDIDATES = [
    ("/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
     "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Bold.ttf",
     "Deva", "Deva-Bold"),
    (_WIN + "/NotoSansDevanagari-Regular.ttf",
     _WIN + "/NotoSansDevanagari-Bold.ttf",
     "Deva", "Deva-Bold"),
    (_WIN + "/ARIALUNI.TTF", _WIN + "/ARIALUNI.TTF", "ArialUni", "ArialUni"),
    (_WIN + "/mangal.ttf",   _WIN + "/mangalb.ttf",  "Mangal",   "MangalBold"),
    ("/usr/share/fonts/truetype/freefont/FreeSans.ttf",
     "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
     "FreeSans", "FreeSansBold"),
]

PDF_FONT, PDF_BOLD = "Helvetica", "Helvetica-Bold"
HINDI_FONT_PATH = None

_HINDI_IMG_CANDIDATES = [
    "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
    _WIN + "/NotoSansDevanagari-Regular.ttf",
    _WIN + "/ARIALUNI.TTF",
    _WIN + "/mangal.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
]
for _hfp in _HINDI_IMG_CANDIDATES:
    if os.path.exists(_hfp):
        HINDI_FONT_PATH = _hfp
        break

for _r, _b, _name, _bname in _FONT_CANDIDATES:
    if os.path.exists(_r) and os.path.exists(_b):
        try:
            pdfmetrics.registerFont(TTFont(_name,  _r))
            pdfmetrics.registerFont(TTFont(_bname, _b))
            PDF_FONT, PDF_BOLD = _name, _bname
            print(f"PDF font: {_name}")
            break
        except Exception:
            continue

if PDF_FONT == "Helvetica":
    try:
        _base = "https://github.com/googlefonts/noto-fonts/raw/main/hinted/ttf/NotoSansDevanagari"
        _tmp  = tempfile.mkdtemp()
        _dr   = os.path.join(_tmp, "NotoSansDevanagari-Regular.ttf")
        _db   = os.path.join(_tmp, "NotoSansDevanagari-Bold.ttf")
        print("Downloading NotoSansDevanagari font...")
        urllib.request.urlretrieve(f"{_base}/NotoSansDevanagari-Regular.ttf", _dr)
        urllib.request.urlretrieve(f"{_base}/NotoSansDevanagari-Bold.ttf",    _db)
        pdfmetrics.registerFont(TTFont("Deva",      _dr))
        pdfmetrics.registerFont(TTFont("Deva-Bold", _db))
        PDF_FONT, PDF_BOLD = "Deva", "Deva-Bold"
        if HINDI_FONT_PATH is None:
            HINDI_FONT_PATH = _dr
        print("PDF font: Deva (downloaded)")
    except Exception as _fe:
        print(f"WARNING: Could not load Devanagari font ({_fe}). Using Helvetica.")

_LAT  = "Helvetica"
_LATB = "Helvetica-Bold"

# ═══════════════════════════════════════════════════════════════════════════════
# HALKA → KARMCHARI MAPPING
# ═══════════════════════════════════════════════════════════════════════════════
HALKA_MAPPING = {
    "NADAUL":                  "SANJAY KUMAR kc7 new",
    "BERRA":                   "SANJAY KUMAR kc7 new",
    "NAGAR PARISHAD MASAURHI": "SANJAY KUMAR kc7 new",
    "BHADAURA":                "RAJESH kc3 new",
    "DEWARIYA":                "RAJESH kc3 new",
    "LAKHNOUR BEDOULI":        "RAJESH kc3 new",
    "NURA":                    "RAJESH kc3 new",
    "CHAPAUR":                 "DEEPAK kc1 new",
    "KARAY":                   "DEEPAK kc1 new",
    "KHARRAT":                 "DEEPAK kc1 new",
    "NISHIYAWAN":              "DEEPAK kc1 new",
    "BARA":                    "NISHIKANT kc2 new",
    "DAULATPUR":               "NISHIKANT kc2 new",
    "BHAGWANGANJ":             "NISHIKANT kc2 new",
    "TINERI":                  "NISHIKANT kc2 new",
    "CHARMA":                  "VIKAS kc6 new",
    "BHAISHWAN":               "VIKAS kc6 new",
    "REWAN":                   "VIKAS kc6 new",
    "SAHABAD":                 "VIKAS kc6 new",
}
ALL_KARMCHARIS = list(dict.fromkeys(HALKA_MAPPING.values()))

# ═══════════════════════════════════════════════════════════════════════════════
# REPORT CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════════════════════
REPORT_CONFIGS = [
    {
        "report":             "e-Mutation (Defect Check)",
        "bucket":             "10",
        "bucket_color":       ("FCE4D6", "000000"),
        "col_fallback":       10,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "eMutation_DefectCheck_10days",
        "pending_label":      "10",
    },
    {
        "report":             "e-Mutation (Defect Check)",
        "bucket":             "15",
        "bucket_color":       ("FCE4D6", "000000"),
        "col_fallback":       10,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "eMutation_DefectCheck_15days",
        "pending_label":      "15",
    },
    {
        "report":             "e-Mutation (Defect Check)",
        "bucket":             "20",
        "bucket_color":       ("FCE4D6", "000000"),
        "col_fallback":       10,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "eMutation_DefectCheck_20days",
        "pending_label":      "20",
    },
    {
        "report":             "Parimarjan (Digitization of Jamabandi not available Online)",
        "bucket":             "35",
        "bucket_color":       ("FF8C00", "FFFFFF"),
        "col_fallback":       10,
        "skip_cols":          (5, 6),
        "scraper_type":       "parimarjan",
        "aging_extra":        True,
        "remark_filter":      False,
        "file_tag":           "Digitization_35days",
        "pending_label":      "35",
    },
    {
        "report":             "Parimarjan (Digitization of Jamabandi not available Online)",
        "bucket":             "75",
        "bucket_color":       ("E64A19", "FFFFFF"),
        "col_fallback":       11,
        "skip_cols":          (5, 6),
        "scraper_type":       "parimarjan",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "Digitization_75days",
        "pending_label":      "75",
    },
    {
        "report":             "Parimarjan (Digitization of Jamabandi not available Online)",
        "bucket":             "120",
        "bucket_color":       ("B71C1C", "FFFFFF"),
        "col_fallback":       12,
        "skip_cols":          (5, 6),
        "scraper_type":       "parimarjan",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "Digitization_120days",
        "pending_label":      "120",
    },
    {
        "report":             "Parimarjan (Rectification in digitized Jamabandi)",
        "bucket":             "35",
        "bucket_color":       ("FF8C00", "FFFFFF"),
        "col_fallback":       10,
        "skip_cols":          (5, 6),
        "scraper_type":       "rectification",
        "aging_extra":        True,
        "remark_filter":      False,
        "file_tag":           "Rectification_35days",
        "pending_label":      "35",
    },
    {
        "report":             "Parimarjan (Rectification in digitized Jamabandi)",
        "bucket":             "75",
        "bucket_color":       ("E64A19", "FFFFFF"),
        "col_fallback":       11,
        "skip_cols":          (5, 6),
        "scraper_type":       "rectification",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "Rectification_75days",
        "pending_label":      "75",
    },
    {
        "report":             "Parimarjan (Rectification in digitized Jamabandi)",
        "bucket":             "120",
        "bucket_color":       ("B71C1C", "FFFFFF"),
        "col_fallback":       12,
        "skip_cols":          (5, 6),
        "scraper_type":       "rectification",
        "aging_extra":        False,
        "remark_filter":      True,
        "file_tag":           "Rectification_120days",
        "pending_label":      "120",
    },
    {
        "report":             "e-Mutation",
        "bucket":             "35",
        "bucket_color":       ("FF8C00", "FFFFFF"),
        "col_fallback":       10,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        True,
        "remark_filter":      False,
        "file_tag":           "eMutation_35days",
        "pending_label":      "35",
    },
    {
        "report":             "e-Mutation",
        "bucket":             "75",
        "bucket_color":       ("E64A19", "FFFFFF"),
        "col_fallback":       11,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        False,
        "remark_filter":      False,
        "file_tag":           "eMutation_75days",
        "pending_label":      "75",
    },
    {
        "report":             "e-Mutation",
        "bucket":             "120",
        "bucket_color":       ("B71C1C", "FFFFFF"),
        "col_fallback":       12,
        "skip_cols":          (3,),
        "scraper_type":       "emutation",
        "aging_extra":        False,
        "remark_filter":      False,
        "file_tag":           "eMutation_120days",
        "pending_label":      "120",
    },
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHARED REMARK EXCLUSION CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
EXCLUDED_EXACT    = {"--", "Application Disposed", "nan"}
EXCLUDED_PREFIXES = ("Reverted to Applicant",)

_UPLOAD_EXTENSIONS = (".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx", ".zip", ".rar")
_VALID_APPID = re.compile(r'^\d{15}$')

# ═══════════════════════════════════════════════════════════════════════════════
# PERFORMANCE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
VIEW_CONCURRENCY  = 16
VIEW_GOTO_TIMEOUT = 30_000
VIEW_WAIT_TIMEOUT = 20_000
VIEW_POLL_MS      = 200
VIEW_POLL_MAX     = 150
MAX_OPEN_TABS     = 20

EARLY_EXIT_STABLE_POLLS = 8

# ═══════════════════════════════════════════════════════════════════════════════
# JAVASCRIPT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
_JS_TABLE_READY_AND_EXTRACT = r"""
() => {
    function looksLikeDate(s) {
        return /\d{2}[\/\-]\d{2}[\/\-]\d{4}|\d{4}[\/\-]\d{2}[\/\-]\d{2}/.test(s);
    }
    function isPagination(tbl) {
        const trs = [...tbl.querySelectorAll('tr')];
        if (trs.length === 0) return true;
        const firstCells = [...trs[0].querySelectorAll('th,td')]
            .map(c => c.innerText.trim()).filter(t => t.length > 0);
        if (firstCells.length === 0) return true;
        return firstCells.every(t => /^\d{1,3}$/.test(t));
    }
    function hasDateData(tbl) {
        const trs = [...tbl.querySelectorAll('tr')];
        return trs.slice(1).some(tr =>
            [...tr.querySelectorAll('td')].some(td => looksLikeDate(td.innerText.trim()))
        );
    }
    function isDocumentTable(tbl) {
        const trs = [...tbl.querySelectorAll('tr')];
        if (trs.length === 0) return false;
        const hdr = [...trs[0].querySelectorAll('th,td')]
                        .map(c => c.innerText.toLowerCase()).join(' ');
        return hdr.includes('document type') || hdr.includes('document no')
            || hdr.includes('issuing authority') || hdr.includes('uploaded document')
            || hdr.includes('view document') || hdr.includes('dar kewla')
            || hdr.includes('\u0926\u0938\u094d\u0924\u093e\u0935\u0947\u091c');
    }
    let actionTbl = null;
    try {
        const xTbl = document.evaluate(
            '/html/body/form/div[3]/div[2]/div[2]/div/table',
            document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null
        ).singleNodeValue;
        if (xTbl && hasDateData(xTbl) && !isDocumentTable(xTbl)) actionTbl = xTbl;
    } catch(e) {}
    if (!actionTbl) {
        try {
            const xTbl2 = document.evaluate(
                '/html/body/form/div[3]/div[4]/div[2]/div/table',
                document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null
            ).singleNodeValue;
            if (xTbl2 && hasDateData(xTbl2) && !isDocumentTable(xTbl2)) actionTbl = xTbl2;
        } catch(e) {}
    }
    if (!actionTbl) {
        const allTbls = [...document.querySelectorAll('table')];
        for (const tbl of allTbls) {
            if (isPagination(tbl) || isDocumentTable(tbl)) continue;
            const trs = [...tbl.querySelectorAll('tr')];
            if (trs.length < 2) continue;
            const hdr = [...trs[0].querySelectorAll('th,td')]
                            .map(c => c.innerText.toLowerCase()).join(' ');
            const hasDateHdr = hdr.includes('date') || hdr.includes('\u0926\u093f\u0928\u093e\u0902\u0915');
            const hasActionHdr = hdr.includes('remark') || hdr.includes(' by')
                              || hdr.includes('action')
                              || hdr.includes('\u091f\u093f\u092a\u094d\u092a\u0923\u0940')
                              || hdr.includes('\u0926\u094d\u0935\u093e\u0930\u093e')
                              || hdr.includes('\u0915\u093e\u0930\u094d\u0930\u0935\u093e\u0908');
            if (hasDateHdr && hasActionHdr && hasDateData(tbl)) {
                actionTbl = tbl; break;
            }
        }
    }
    if (!actionTbl) {
        const allTbls = [...document.querySelectorAll('table')].reverse();
        for (const tbl of allTbls) {
            if (isPagination(tbl) || isDocumentTable(tbl)) continue;
            const trs = [...tbl.querySelectorAll('tr')];
            if (trs.length < 2) continue;
            if (hasDateData(tbl)) { actionTbl = tbl; break; }
        }
    }
    if (!actionTbl) return null;
    const trs = [...actionTbl.querySelectorAll('tr')];
    if (trs.length < 2) return null;
    const hasData = [...trs].slice(1).some(tr =>
        [...tr.querySelectorAll('td')].some(td => td.innerText.trim().length > 0));
    if (!hasData) return null;
    let dateIdx = 0, remarkIdx = -1, byIdx = -1;
    const hdrCells = [...trs[0].querySelectorAll('th,td')].map(c => c.innerText.toLowerCase());
    const totalCols = hdrCells.length;
    hdrCells.forEach((txt, i) => {
        if (txt.includes('date') || txt.includes('\u0926\u093f\u0928\u093e\u0902\u0915'))
            dateIdx = i;
        if (txt.includes('remark') || txt.includes('\u091f\u093f\u092a\u094d\u092a\u0923\u0940')
         || txt.includes('\u0915\u093e\u0930\u094d\u0930\u0935\u093e\u0908'))
            remarkIdx = i;
        if (txt.includes('taken by') || txt.includes(' by')
         || txt.includes('\u0926\u094d\u0935\u093e\u0930\u093e')
         || txt.includes('\u0905\u0927\u093f\u0915\u093e\u0930\u0940'))
            byIdx = i;
    });
    if (totalCols === 3) {
        if (remarkIdx === -1) remarkIdx = 1;
        if (byIdx     === -1) byIdx     = 2;
    } else if (totalCols === 2) {
        if (byIdx === -1) byIdx = 1;
    } else if (totalCols > 3) {
        if (remarkIdx === -1) remarkIdx = totalCols - 2;
        if (byIdx     === -1) byIdx     = totalCols - 1;
    }
    if (remarkIdx !== -1 && remarkIdx === byIdx) {
        byIdx = remarkIdx + 1 < totalCols ? remarkIdx + 1 : remarkIdx - 1;
        if (byIdx < 0) byIdx = remarkIdx;
    }
    const validRows = [];
    for (let i = 1; i < trs.length; i++) {
        const cells = [...trs[i].querySelectorAll('td')].map(td => td.innerText.trim());
        if (cells.length < 1 || !cells.some(c => c.length > 0)) continue;
        const date   = cells[dateIdx] || '--';
        const remark = remarkIdx >= 0 && remarkIdx < cells.length ? (cells[remarkIdx] || '--') : '--';
        const by     = byIdx     >= 0 && byIdx     < cells.length ? (cells[byIdx]     || '--') : '--';
        if (!looksLikeDate(date)) continue;
        validRows.push({ date, remark, by });
    }
    if (validRows.length === 0) return null;
    return { rows: validRows, ready: true };
}
"""

_JS_PAGE_SETTLED = r"""
() => { return document.readyState === 'complete'; }
"""

_JS_ACTIVE_PAGER_PAGE = r"""
(pnum) => {
    // ASP.NET GridView renders the currently-active page as a <span> (non-clickable)
    // inside a <tr> that contains only small digit tokens — the pager row.
    // We scope our search to pager rows only to avoid false positives from data cells
    // (e.g. application IDs that happen to contain the same digit sequence).

    const pnum_str = String(pnum);

    // Find candidate rows that look like a GridView pager row:
    // every non-empty cell is a short number, "..." or a link with a short number.
    function isPagerRow(tr) {
        const cells = [...tr.querySelectorAll('td,th')];
        if (cells.length === 0) return false;
        const texts = cells.map(c => c.innerText.trim()).filter(t => t.length > 0);
        return texts.length > 0 && texts.every(t => /^[\d.]+$/.test(t) || /^\d{1,3}$/.test(t));
    }

    // Collect all pager rows
    const pagerRows = [...document.querySelectorAll('tr')].filter(isPagerRow);

    if (pagerRows.length === 0) {
        // Fallback: any <span> with exact text pnum and no sibling <a> with same text
        const spanMatch = [...document.querySelectorAll('span')]
            .some(s => s.innerText.trim() === pnum_str);
        const linkMatch = [...document.querySelectorAll('a')]
            .some(a => a.innerText.trim() === pnum_str);
        return spanMatch && !linkMatch;
    }

    // In pager rows: span with pnum text exists AND no link with pnum text exists
    const spanInPager = pagerRows.some(tr =>
        [...tr.querySelectorAll('span')].some(s => s.innerText.trim() === pnum_str)
    );
    const linkInPager = pagerRows.some(tr =>
        [...tr.querySelectorAll('a')].some(a => a.innerText.trim() === pnum_str)
    );

    return spanInPager && !linkInPager;
}
"""

_JS_GET_VIEW_PAGE_LINKS = r"""
() => {
    const results = [];
    const seen = new Set();
    [...document.querySelectorAll('a')].forEach(a => {
        const txt = a.innerText.trim();
        if (/^\d+$/.test(txt) && !seen.has(txt)) {
            seen.add(txt);
            results.push({ text: txt, type: 'a', href: a.href || '',
                           onclick: a.getAttribute('onclick') || '', id: a.id || '' });
        }
    });
    [...document.querySelectorAll('input[type=submit],input[type=button]')].forEach(inp => {
        const txt = (inp.value || '').trim();
        if (/^\d+$/.test(txt) && !seen.has(txt)) {
            seen.add(txt);
            results.push({ text: txt, type: 'input', href: '',
                           onclick: inp.getAttribute('onclick') || '',
                           id: inp.id || '', name: inp.name || '' });
        }
    });
    [...document.querySelectorAll('[onclick]')].forEach(el => {
        const oc = el.getAttribute('onclick') || '';
        const txt = (el.innerText || el.value || '').trim();
        if (/^\d+$/.test(txt) && oc.includes('doPostBack') && !seen.has(txt)) {
            seen.add(txt);
            results.push({ text: txt, type: 'postback', href: '',
                           onclick: oc, id: el.id || '' });
        }
    });
    return results;
}
"""

# ── FIXED: Ellipsis click handler for ASP.NET GridView pagers ───────────────────
#
# Root cause (visible in screenshot): the BiharBhumi portal uses ASP.NET
# GridView with a pager that shows "1 2 3 4 5 6 7 8 9 10 ..." where the
# visible "..." text is rendered as a <span> element (non-clickable), NOT as
# an <a> tag.  The actual postback trigger for jumping to the next page-number
# group is a sibling or parent <a> element whose onclick contains __doPostBack
# with a "Page$..." argument pattern.
#
# Strategy (3-tier fallback):
#   Tier 1 — Direct <a> with text "..." or "…": works when the portal
#             renders the ellipsis as a clickable link (some ASP.NET versions).
#   Tier 2 — <span> text "...": walk up to find the nearest ancestor <a>
#             or sibling <a> with __doPostBack onclick; click that.
#   Tier 3 — __doPostBack injection: scan all existing pager <a> onclick
#             values to find the largest visible page number N, then infer
#             that the "..." postback argument is "Page$" + (N+1) and call
#             __doPostBack directly.  This is the most robust fallback.
_JS_CLICK_ELLIPSIS = r"""
() => {
    // ── Tier 1: <a> whose trimmed text is exactly "..." or the Unicode ellipsis ──
    const directLinks = [...document.querySelectorAll('a')].filter(a => {
        const t = a.innerText.trim();
        return t === '...' || t === '\u2026' || /^\.{3,}$/.test(t);
    });
    if (directLinks.length > 0) {
        directLinks[directLinks.length - 1].click();
        return { tier: 1 };
    }

    // ── Tier 2: <span> with "..." text — find nearest postback-carrying <a> ────
    const ellipsisSpans = [...document.querySelectorAll('span')].filter(s => {
        const t = s.innerText.trim();
        return t === '...' || t === '\u2026' || /^\.{3,}$/.test(t);
    });
    for (const span of ellipsisSpans.reverse()) {
        // Check parent chain up to 4 levels
        let node = span;
        for (let i = 0; i < 4; i++) {
            node = node.parentElement;
            if (!node) break;
            if (node.tagName === 'A' && (node.getAttribute('onclick') || '').includes('doPostBack')) {
                node.click();
                return { tier: 2, via: 'parent' };
            }
        }
        // Check siblings
        if (span.parentElement) {
            const siblings = [...span.parentElement.children];
            const idx = siblings.indexOf(span);
            // Look at siblings after this span
            for (let i = idx + 1; i < siblings.length; i++) {
                const sib = siblings[i];
                if (sib.tagName === 'A' && (sib.getAttribute('onclick') || '').includes('doPostBack')) {
                    sib.click();
                    return { tier: 2, via: 'sibling-after' };
                }
            }
            // Look at siblings before this span
            for (let i = idx - 1; i >= 0; i--) {
                const sib = siblings[i];
                if (sib.tagName === 'A' && (sib.getAttribute('onclick') || '').includes('doPostBack')) {
                    sib.click();
                    return { tier: 2, via: 'sibling-before' };
                }
            }
        }
    }

    // ── Tier 3: __doPostBack injection ──────────────────────────────────────────
    // Find the largest page number currently visible as a pager link.
    // The "..." jumps to the group starting at (largestVisible + 1).
    let maxPageNum = 0;
    let postbackTarget = null;

    [...document.querySelectorAll('a[onclick*="doPostBack"]')].forEach(a => {
        const oc = a.getAttribute('onclick') || '';
        // ASP.NET pattern: __doPostBack('GridView1','Page$5')
        const m = oc.match(/__doPostBack\(['"]([^'"]+)['"]\s*,\s*['"]Page\$(\d+)['"]\)/);
        if (m) {
            const pnum = parseInt(m[2], 10);
            if (pnum > maxPageNum) {
                maxPageNum = pnum;
                postbackTarget = m[1];
            }
        }
    });

    if (postbackTarget !== null && maxPageNum > 0) {
        const nextGroup = maxPageNum + 1;
        try {
            __doPostBack(postbackTarget, 'Page$' + nextGroup);
            return { tier: 3, target: postbackTarget, nextGroup };
        } catch(e) {
            // If __doPostBack is namespaced (WebForm_DoPostBack etc.), try eval
            try {
                const allScripts = document.documentElement.innerHTML;
                const fnMatch = allScripts.match(/function\s+(WebForm_DoPostBack|theForm\.submit)\b/);
                return { tier: 3, failed: true, error: String(e) };
            } catch(e2) {
                return { tier: 3, failed: true, error: String(e2) };
            }
        }
    }

    return null;  // No ellipsis found by any method
}
"""

_JS_EXTRACT_ROWS_EMUTATION = r"""
() => {
    function isPaginationRow(cells) {
        const nonEmpty = cells.filter(c => c.length > 0);
        if (nonEmpty.length === 0) return true;
        if (nonEmpty.every(c => /^\d{1,3}$/.test(c))) return true;
        if (nonEmpty.every(c => /^[\d\s.\u2026]+$/.test(c))) return true;
        return false;
    }
    function isPaginationTable(tbl) {
        const trs = [...tbl.querySelectorAll('tr')];
        if (trs.length === 0) return true;
        const firstCells = [...trs[0].querySelectorAll('th,td')]
            .map(c => c.innerText.trim()).filter(t => t.length > 0);
        return firstCells.length > 0 && firstCells.every(t => /^\d{1,3}$/.test(t));
    }
    const tables = [...document.querySelectorAll('table')];
    const tbl = tables.filter(t => !isPaginationTable(t))
        .reduce((b,t) =>
            t.querySelectorAll('tr').length > (b ? b.querySelectorAll('tr').length : 0) ? t : b, null);
    if (!tbl) return { headers:[], rows:[] };
    const allTrs = [...tbl.querySelectorAll('tr')];
    let hdrIdx = 0;
    for (let i = 0; i < Math.min(10, allTrs.length); i++)
        if (allTrs[i].querySelectorAll('th').length > 0) { hdrIdx = i; break; }
    const headers = [...allTrs[hdrIdx].querySelectorAll('th')].map(c=>c.innerText.trim()).filter(Boolean);
    const rows = [];
    for (let ri = hdrIdx+1; ri < allTrs.length; ri++) {
        const tds = [...allTrs[ri].querySelectorAll('td')];
        if (tds.length < 2) continue;
        const cells = tds.map(td=>td.innerText.trim());
        if (isPaginationRow(cells)) continue;
        if (/total|\u0915\u0941\u0932|grand|sum|\u092f\u094b\u0917/i.test(cells[0]||'')) continue;
        if (cells.filter(c=>c).length < 2) continue;
        const hasTokenLike = cells.some(c => /[A-Z]/.test(c) || c.includes('/'));
        const allShortDigits = cells.filter(c=>c).every(c => /^\d{1,4}$/.test(c));
        if (!hasTokenLike && allShortDigits) continue;
        let url = '';
        const btns = [...allTrs[ri].querySelectorAll('a,input,button')];
        for (const b of btns) {
            const txt = (b.innerText||b.value||'').toLowerCase();
            if (txt.includes('view')||txt.includes('\u0926\u0947\u0916\u0947\u0902')||txt.includes('\u0926\u0947\u0916\u0947')) {
                const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
                if (m) { url=m[1]; break; }
                if (b.href&&b.href.startsWith('http')&&!b.href.includes('javascript')) { url=b.href; break; }
            }
        }
        if (!url && btns.length) {
            const b = btns[0];
            const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
            if (m) url=m[1];
            else if (b.href&&b.href.startsWith('http')&&!b.href.includes('javascript')) url=b.href;
        }
        rows.push({cells, url, trIdx:ri});
    }
    return {headers, rows};
}
"""

_JS_TRY_EXTRACT_ACTION = r"""
() => {
    let container = null;
    const xpaths = [
        '/html/body/form/div[3]/div/div/div/div/div/div[4]/div[2]',
        '/html/body/form/div[3]/div/div/div/div/div/div[3]/div[2]',
        '/html/body/form/div[3]/div/div/div/div/div/div[5]/div[2]',
        '/html/body/form/div[3]/div/div/div/div/div/div[4]/div[1]',
    ];
    for (const xp of xpaths) {
        try {
            const node = document.evaluate(
                xp, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null
            ).singleNodeValue;
            if (node && node.querySelector('table')) { container = node; break; }
        } catch(e) {}
    }
    if (!container) {
        const allTables = Array.from(document.querySelectorAll('table'));
        for (const t of allTables) {
            const hdrs = Array.from(t.querySelectorAll('th, tr:first-child td'))
                              .map(c => c.innerText.toLowerCase()).join(' ');
            if ((hdrs.includes('date') || hdrs.includes('\u0926\u093f\u0928\u093e\u0902\u0915')) &&
                (hdrs.includes('remark') || hdrs.includes('by') ||
                 hdrs.includes('\u091f\u093f\u092a\u094d\u092a\u0923\u0940') ||
                 hdrs.includes('\u0926\u094d\u0935\u093e\u0930\u093e'))) {
                container = t.parentElement || t; break;
            }
        }
    }
    if (!container) return null;
    const tbl = (container.tagName === 'TABLE') ? container : container.querySelector('table');
    if (!tbl) return null;
    const trs = Array.from(tbl.querySelectorAll('tr'));
    if (trs.length < 2) return null;
    let dateIdx = 0, remarkIdx = 1, byIdx = 2;
    const hdrCells = Array.from(trs[0].querySelectorAll('th, td'))
                          .map(c => c.innerText.toLowerCase());
    hdrCells.forEach((txt, i) => {
        if (txt.includes('date') || txt.includes('\u0926\u093f\u0928\u093e\u0902\u0915'))   dateIdx   = i;
        if (txt.includes('remark') || txt.includes('\u091f\u093f\u092a\u094d\u092a\u0923\u0940')) remarkIdx = i;
        if (txt.includes('by') || txt.includes('\u0926\u094d\u0935\u093e\u0930\u093e'))      byIdx     = i;
    });
    for (let i = trs.length - 1; i >= 1; i--) {
        const cells = Array.from(trs[i].querySelectorAll('td'))
                           .map(td => td.innerText.trim());
        if (cells.length < 2) continue;
        return {
            date:   cells[dateIdx]   || '--',
            remark: cells[remarkIdx] || '--',
            by:     cells[byIdx]     || '--'
        };
    }
    return null;
}
"""

_JS_EXTRACT_LAND_DETAILS = r"""
() => {
    function norm(s) { return (s || '').trim(); }
    function isPagination(tbl) {
        const trs = [...tbl.querySelectorAll('tr')];
        if (trs.length === 0) return true;
        const firstCells = [...trs[0].querySelectorAll('th,td')]
            .map(c => norm(c.innerText)).filter(t => t.length > 0);
        if (firstCells.length === 0) return true;
        return firstCells.every(t => /^\d{1,3}$/.test(t));
    }

    const allTbls = [...document.querySelectorAll('table')];
    let landTbl = null;

    for (const tbl of allTbls) {
        if (isPagination(tbl)) continue;
        const trs = [...tbl.querySelectorAll('tr')];
        if (trs.length === 0) continue;
        const hdrTexts = [...trs[0].querySelectorAll('th,td')]
            .map(c => norm(c.innerText).toLowerCase());
        const hdrJoined = hdrTexts.join(' | ');
        const hasHalka      = hdrJoined.includes('halka');
        const hasMauja       = hdrJoined.includes('mauja');
        const hasKhataOrPlot = hdrJoined.includes('khata') || hdrJoined.includes('plot');
        const isDocTable     = hdrJoined.includes('document') || hdrJoined.includes('issuing authority');
        if (hasHalka && hasMauja && hasKhataOrPlot && !isDocTable) { landTbl = tbl; break; }
    }
    if (!landTbl) return null;

    const trs = [...landTbl.querySelectorAll('tr')];
    const hdrTexts = [...trs[0].querySelectorAll('th,td')].map(c => norm(c.innerText).toLowerCase());

    let iHalka = -1, iMauja = -1, iThana = -1, iKhata = -1, iPlot = -1, iArea1 = -1, iArea2 = -1;
    hdrTexts.forEach((txt, i) => {
        if (iHalka === -1 && txt.includes('halka'))                                              iHalka = i;
        if (iMauja === -1 && txt.includes('mauja'))                                               iMauja = i;
        if (iThana === -1 && txt.includes('thana'))                                               iThana = i;
        if (iKhata === -1 && txt.includes('khata'))                                               iKhata = i;
        if (iPlot  === -1 && txt.includes('plot'))                                                iPlot  = i;
        if (iArea1 === -1 && txt.includes('area') && (txt.includes('acre') || txt.includes('1')))    iArea1 = i;
        if (iArea2 === -1 && txt.includes('area') && (txt.includes('decimal') || txt.includes('2'))) iArea2 = i;
    });

    function pick(cells, i) { return (i >= 0 && i < cells.length && cells[i]) ? cells[i] : '--'; }
    const out = { halka: [], mauja: [], thana: [], khataNo: [], plotNo: [], area1: [], area2: [] };
    for (let r = 1; r < trs.length; r++) {
        const cells = [...trs[r].querySelectorAll('td')].map(td => norm(td.innerText));
        if (cells.length === 0 || cells.every(c => c.length === 0)) continue;
        out.halka.push(pick(cells, iHalka));
        out.mauja.push(pick(cells, iMauja));
        out.thana.push(pick(cells, iThana));
        out.khataNo.push(pick(cells, iKhata));
        out.plotNo.push(pick(cells, iPlot));
        out.area1.push(pick(cells, iArea1));
        out.area2.push(pick(cells, iArea2));
    }
    if (out.halka.length === 0) return null;
    return out;
}
"""

_JS_EXTRACT_ROWS_PARIMARJAN = r"""
() => {
    const tables = [...document.querySelectorAll('table')];
    const tbl = tables.reduce((b, t) =>
        t.querySelectorAll('tr').length > (b ? b.querySelectorAll('tr').length : 0) ? t : b, null);
    if (!tbl) return { headers: [], rows: [] };
    const allTrs = [...tbl.querySelectorAll('tr')];
    let hdrIdx = 0;
    for (let i = 0; i < Math.min(10, allTrs.length); i++) {
        if (allTrs[i].querySelectorAll('th').length > 0) { hdrIdx = i; break; }
    }
    const headers = [...allTrs[hdrIdx].querySelectorAll('th')]
                        .map(c => c.innerText.trim()).filter(Boolean);
    const rows = [];
    for (let ri = hdrIdx + 1; ri < allTrs.length; ri++) {
        const tds = [...allTrs[ri].querySelectorAll('td')];
        if (tds.length < 2) continue;
        const cells = tds.map(td => td.innerText.trim());
        if (/total|\u0915\u0941\u0932|grand|sum|\u092f\u094b\u0917/i.test(cells[0]||'')) continue;
        if (cells.filter(c => c).length < 2) continue;
        let url = '';
        const btns = [...allTrs[ri].querySelectorAll('a,input,button')];
        for (const b of btns) {
            const txt = (b.innerText || b.value || '').toLowerCase();
            if (txt.includes('view') || txt.includes('\u0926\u0947\u0916\u0947\u0902') || txt.includes('\u0926\u0947\u0916\u0947')) {
                const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
                if (m) { url = m[1]; break; }
                if (b.href && b.href.startsWith('http') && !b.href.includes('javascript'))
                    { url = b.href; break; }
            }
        }
        if (!url && btns.length) {
            const b = btns[0];
            const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
            if (m) url = m[1];
            else if (b.href && b.href.startsWith('http') && !b.href.includes('javascript'))
                url = b.href;
        }
        rows.push({ cells, url, trIdx: ri });
    }
    return { headers, rows };
}
"""

_JS_EXTRACT_ROWS_RECTIFICATION = r"""
() => {
    const tables = [...document.querySelectorAll('table')];
    const tbl = tables.reduce((b, t) =>
        t.querySelectorAll('tr').length > (b ? b.querySelectorAll('tr').length : 0) ? t : b, null);
    if (!tbl) return { headers: [], rows: [] };
    const allTrs = [...tbl.querySelectorAll('tr')];
    let hdrIdx = 0;
    for (let i = 0; i < Math.min(10, allTrs.length); i++) {
        if (allTrs[i].querySelectorAll('th').length > 0) { hdrIdx = i; break; }
    }
    const headers = [...allTrs[hdrIdx].querySelectorAll('th')]
                        .map(c => c.innerText.trim()).filter(Boolean);
    const rows = [];
    for (let ri = hdrIdx + 1; ri < allTrs.length; ri++) {
        const tds = [...allTrs[ri].querySelectorAll('td')];
        if (tds.length < 2) continue;
        const cells = tds.map(td => td.innerText.trim());
        if (/total|\u0915\u0941\u0932|grand|sum|\u092f\u094b\u0917/i.test(cells[0]||'')) continue;
        if (cells.filter(c => c).length < 2) continue;
        const v0 = parseInt(cells[0]);
        const v1 = parseInt(cells[1]);
        if (!isNaN(v0) && !isNaN(v1) && v0 >= 1 && v0 <= 9 && v1 >= 1 && v1 <= 9) continue;
        const appIdCol = 1;
        if (cells.length > appIdCol && !/^\d{15}$/.test(cells[appIdCol])) continue;
        let url = '';
        const btns = [...allTrs[ri].querySelectorAll('a,input,button')];
        for (const b of btns) {
            const txt = (b.innerText || b.value || '').toLowerCase();
            if (txt.includes('view') || txt.includes('\u0926\u0947\u0916\u0947\u0902') || txt.includes('\u0926\u0947\u0916\u0947')) {
                const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
                if (m) { url = m[1]; break; }
                if (b.href && b.href.startsWith('http') && !b.href.includes('javascript'))
                    { url = b.href; break; }
            }
        }
        if (!url && btns.length) {
            const b = btns[0];
            const m = (b.getAttribute('onclick')||'').match(/window\.open\(['"]([^'"]+)['"]/);
            if (m) url = m[1];
            else if (b.href && b.href.startsWith('http') && !b.href.includes('javascript'))
                url = b.href;
        }
        rows.push({ cells, url, trIdx: ri });
    }
    return { headers, rows };
}
"""


# ═══════════════════════════════════════════════════════════════════════════════
# BROWSER / PAGE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
async def dismiss_dialogs(page):
    try: page.remove_all_listeners("dialog")
    except: pass
    page.on("dialog", lambda d: asyncio.create_task(d.accept()))


# ── Session expiry detection + auto re-login ────────────────────────────────
async def is_session_expired(pg) -> bool:
    """Return True if the page looks like a logout/session-expired page."""
    try:
        url   = pg.url.lower()
        title = (await pg.title()).lower()
        body  = (await pg.evaluate(
            "document.body ? document.body.innerText : ''")).lower()
    except Exception:
        return False

    logout_signals = [
        "logout" in url,
        "loggedout" in url,
        "you have been logged out" in body,
        "session expired" in body,
        "session has expired" in body,
        "please login" in body and "login" in url,
        "log out" in title,
    ]
    return any(logout_signals)


async def re_login(context, page) -> "Page":
    """
    Detect session expiry on any open page and re-login.
    Returns the new dashboard page (or the original page if not expired).
    Raises RuntimeError after 3 failed attempts.
    """
    # Check all open pages for logout indicators
    expired_page = None
    for pg in list(context.pages):
        try:
            if await is_session_expired(pg):
                expired_page = pg
                break
        except Exception:
            continue

    if expired_page is None:
        return page   # session still valid

    print("\n⚠️  SESSION EXPIRED — attempting re-login...")

    for attempt in range(3):
        try:
            print(f"   Re-login attempt {attempt + 1}...")
            # Navigate login page in the expired page (or page arg)
            login_pg = expired_page
            await login_pg.goto(
                "https://biharbhumi.bihar.gov.in/BiharEmutationNew/",
                wait_until="load", timeout=90_000)

            await login_pg.select_option("#cbo_District_Code", label="State")
            await login_pg.wait_for_load_state("domcontentloaded")
            await login_pg.select_option("#cbo_Circle_Code1", label="Bihar")
            await login_pg.fill("#txt_UserName", "Bihar")
            await login_pg.fill("#txt_Password1", "nic123")

            n1 = await login_pg.locator("#txtNum1").inner_text()
            n2 = await login_pg.locator("#txtNum2").inner_text()
            await login_pg.fill("#txtResult", str(int(n1.strip()) + int(n2.strip())))
            await dismiss_dialogs(login_pg)

            async with context.expect_page(timeout=30_000) as npi:
                await login_pg.click("#btn_Login")
            new_dash = await npi.value
            await new_dash.wait_for_load_state("load")
            await dismiss_dialogs(new_dash)
            print("   ✅ Re-login successful.")
            return new_dash

        except Exception as e:
            print(f"   Re-login attempt {attempt + 1} failed: {e}")
            await asyncio.sleep(3)

    raise RuntimeError("Re-login failed after 3 attempts.")


async def ensure_session(context, page, dash):
    """
    Check both page and dash for session expiry.
    If expired, re-login and return the new dash.
    Otherwise returns dash unchanged.
    """
    for pg in [dash, page]:
        try:
            if await is_session_expired(pg):
                return await re_login(context, page)
        except Exception:
            pass
    return dash


async def close_strays(context, keep):
    for pg in list(context.pages):
        if pg not in keep:
            try: await pg.close()
            except: pass

async def live_dashboard(context, dash):
    try: await dash.title(); return dash
    except:
        for pg in reversed(context.pages):
            try:
                if "biharbhumi" in pg.url and "detail" not in pg.url.lower():
                    await dismiss_dialogs(pg); return pg
            except: pass
    return context.pages[-1] if context.pages else dash

async def goto_reports(page, xpath):
    for attempt in range(3):
        try:
            await page.click(f"{xpath}/span" if not xpath.endswith("span") else xpath)
            await page.wait_for_load_state("domcontentloaded", timeout=20000)
            await asyncio.sleep(0.5)
            return True
        except Exception as e:
            print(f"  Sidebar attempt {attempt+1}: {e}")
            await asyncio.sleep(1)
    return False

def block_assets(route):
    asyncio.create_task(route.abort())

async def _block_all_except_html(route):
    if route.request.resource_type in ("document", "script", "xhr", "fetch"):
        await route.continue_()
    else:
        await route.abort()


# ═══════════════════════════════════════════════════════════════════════════════
# REMARK HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _is_upload_remark(remark: str) -> bool:
    r = str(remark or "").strip().lower()
    if not r or r == "--":
        return False
    return any(r.endswith(ext) for ext in _UPLOAD_EXTENSIONS)

def _is_real_officer(by: str) -> bool:
    b = str(by or "").strip()
    return b not in ("--", "", "None", "nan")

def _remark_should_exclude(series: pd.Series) -> pd.Series:
    null_mask  = series.isna() | (series.astype(str).str.strip() == "")
    s          = series.fillna("nan").astype(str).str.strip()
    exact_mask = s.str.lower().isin({v.lower() for v in EXCLUDED_EXACT})
    prefix_mask = pd.Series(False, index=series.index)
    for prefix in EXCLUDED_PREFIXES:
        prefix_mask = prefix_mask | s.str.startswith(prefix)
    return null_mask | exact_mask | prefix_mask

def _flatten_land(land) -> tuple:
    if not land:
        return ("--",) * 7

    def j(lst):
        vals = [v for v in (lst or []) if v and v != "--"]
        return "; ".join(vals) if vals else "--"

    return (
        j(land.get("halka")), j(land.get("mauja")), j(land.get("thana")),
        j(land.get("khataNo")), j(land.get("plotNo")),
        j(land.get("area1")), j(land.get("area2")),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PAGINATION HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
async def _click_page_num(pg, pnum: int) -> bool:
    """Click a pager link whose visible text is exactly pnum. Returns True if clicked."""
    link = pg.locator(f"a:text-is('{pnum}')")
    if await link.count() == 0:
        link = pg.locator(f"xpath=//a[normalize-space(text())='{pnum}']")
    if await link.count() > 0:
        await link.first.click(); return True

    inp = pg.locator(f"input[value='{pnum}']")
    if await inp.count() > 0:
        await inp.first.click(); return True

    all_els = await pg.evaluate(f"""() => {{
        return [...document.querySelectorAll('[onclick]')]
            .filter(el => (el.innerText || el.value || '').trim() === '{pnum}')
            .map(el => el.id || el.name || '');
    }}""")
    if all_els and all_els[0]:
        try:
            await pg.locator(f"#{all_els[0]}").first.click(); return True
        except:
            pass

    return False


async def _wait_for_dom_change(pg, pre_html: int, max_polls: int = 80) -> None:
    """Poll until document.body.innerHTML.length differs from pre_html,
    OR until the pager link set changes (handles ASP.NET GridView re-renders
    where the pager group flips but total body byte-length stays the same)."""
    try:
        pre_pager = await pg.evaluate("""() => {
            return [...document.querySelectorAll('a')]
                .map(a => a.innerText.trim())
                .filter(t => /^\\d+$/.test(t))
                .join(',');
        }""")
    except Exception:
        pre_pager = ""

    for _ in range(max_polls):
        await asyncio.sleep(0.1)
        try:
            cur_len = await pg.evaluate("document.body.innerHTML.length")
            if cur_len != pre_html:
                break
            cur_pager = await pg.evaluate("""() => {
                return [...document.querySelectorAll('a')]
                    .map(a => a.innerText.trim())
                    .filter(t => /^\\d+$/.test(t))
                    .join(',');
            }""")
            if cur_pager != pre_pager and cur_pager:
                break
        except Exception:
            break

    await asyncio.sleep(0.3)
    try:
        await pg.wait_for_load_state("domcontentloaded", timeout=5000)
    except Exception:
        pass


async def _click_ellipsis(pg) -> bool:
    """
    Click the ASP.NET pager "..." to reveal the next page-number group.

    Uses _JS_CLICK_ELLIPSIS (3-tier strategy):
      Tier 1 — direct <a> with text "..."
      Tier 2 — <span> "..." → nearest sibling/parent <a> with __doPostBack
      Tier 3 — __doPostBack injection (infer next group from visible pager links)

    Returns True if any tier succeeded (or if Tier 3 was attempted without
    a JS exception — the DOM change wait will verify it actually worked).
    """
    try:
        result = await pg.evaluate(_JS_CLICK_ELLIPSIS)
        if result is None:
            return False
        tier = result.get("tier", 0)
        failed = result.get("failed", False)
        print(f"    [DBG] Ellipsis click — tier {tier}"
              + (f", via={result.get('via')}" if "via" in result else "")
              + (f", nextGroup={result.get('nextGroup')}" if "nextGroup" in result else "")
              + (" FAILED" if failed else ""))
        return not failed
    except Exception as e:
        print(f"    [DBG] Ellipsis JS error: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# VIEW PAGE SCRAPER — e-Mutation style
# ═══════════════════════════════════════════════════════════════════════════════
async def _poll_extract_action_emutation(pg) -> list:
    stable_count  = 0
    last_body_len = -1

    for poll_i in range(VIEW_POLL_MAX):
        try:
            r = await pg.evaluate(_JS_TABLE_READY_AND_EXTRACT)
            if r and r.get("rows"):
                print(f"    [DBG] poll hit at iter {poll_i}: {len(r['rows'])} action rows")
                return r["rows"]
        except Exception as e:
            if poll_i == 0:
                print(f"    [DBG] poll iter {poll_i} eval error: {e}")

        try:
            body_len = await pg.evaluate(
                "document.body ? document.body.innerHTML.length : 0")
        except Exception:
            body_len = last_body_len

        if body_len == last_body_len:
            stable_count += 1
        else:
            stable_count = 0
        last_body_len = body_len

        if stable_count >= EARLY_EXIT_STABLE_POLLS:
            print(f"    [DBG] page stable, no action table after {poll_i+1} polls "
                  f"(~{(poll_i+1)*VIEW_POLL_MS/1000:.1f}s) — bailing early")
            return []

        await asyncio.sleep(VIEW_POLL_MS / 1000)

    try:
        result = await pg.wait_for_function(
            _JS_TABLE_READY_AND_EXTRACT, timeout=5000, polling=50)
        r = await result.json_value()
        if r and r.get("rows"):
            print(f"    [DBG] fallback wait_for_function: {len(r['rows'])} action rows")
            return r["rows"]
    except Exception as e:
        print(f"    [DBG] fallback wait_for_function failed: {e}")

    return []


async def scrape_view_tab_emutation(pg) -> tuple:
    try:
        dbg = await pg.evaluate("""() => ({
            url: window.location.href.slice(0,100),
            bodyLen: document.body ? document.body.innerHTML.length : 0,
            title: document.title,
            tableCount: document.querySelectorAll('table').length
        })""")
        print(f"\n    [DBG] url={dbg['url']} title={dbg['title']} "
              f"bodyLen={dbg['bodyLen']} tables={dbg['tableCount']}")
    except Exception as e:
        print(f"    [DBG] snapshot error: {e}")

    all_action_rows = []

    # ── Page 1 ──────────────────────────────────────────────────────────────────
    rows_p1 = await _poll_extract_action_emutation(pg)
    all_action_rows.extend(rows_p1)
    if not rows_p1:
        print("    [DBG] Page 1: no action rows")

    # ── Forward pagination with ellipsis support ─────────────────────────────────
    current_page = 1
    while True:
        next_page = current_page + 1

        try:
            pre_html = await pg.evaluate("document.body.innerHTML.length")
        except:
            pre_html = 0

        clicked = await _click_page_num(pg, next_page)

        if not clicked:
            print(f"    [DBG] Page {next_page} link not found; trying ellipsis...")
            ellipsis_ok = await _click_ellipsis(pg)

            if not ellipsis_ok:
                print(f"    [DBG] No ellipsis — pagination complete at page {current_page}.")
                break

            print("    Ellipsis triggered — waiting for pager to refresh...")
            await _wait_for_dom_change(pg, pre_html)

            # ASP.NET GridView renders the newly-active page as a <span>, not an <a>.
            # After clicking "...", the portal jumps directly to the first page of the
            # next group (e.g. page 11) and shows that page's data immediately.
            # Check for this case: if next_page is now a span (active), scrape directly.
            # Retry up to 20 times (2s) because the span may appear slightly after
            # _wait_for_dom_change returns.
            is_active_span = False
            for _span_poll in range(20):
                try:
                    is_active_span = await pg.evaluate(_JS_ACTIVE_PAGER_PAGE, next_page)
                except Exception:
                    is_active_span = False
                if is_active_span:
                    break
                await asyncio.sleep(0.1)

            if is_active_span:
                print(f"    [DBG] Page {next_page} is now the active span — "
                      f"data already loaded, scraping directly.")
                # Fall through to scrape below (no extra click needed)
            else:
                # Try clicking the page link in the newly-revealed group
                try:
                    pre_html2 = await pg.evaluate("document.body.innerHTML.length")
                except:
                    pre_html2 = 0

                clicked2 = await _click_page_num(pg, next_page)
                if not clicked2:
                    print(f"    [DBG] Still no page {next_page} after ellipsis — stopping.")
                    break

                await _wait_for_dom_change(pg, pre_html2)
        else:
            await _wait_for_dom_change(pg, pre_html)

        rows_pn = await _poll_extract_action_emutation(pg)
        all_action_rows.extend(rows_pn)
        print(f"    [DBG] Pagination page {next_page}: {len(rows_pn)} action rows")
        current_page = next_page

    # ── 3-priority row selection ─────────────────────────────────────────────────
    if all_action_rows:
        total = len(all_action_rows)
        officer_rows = [r for r in all_action_rows if _is_real_officer(r.get("by", "--"))]
        if officer_rows:
            best = officer_rows[-1]
            print(f"    [DBG] P1 officer (of {total}): {best}")
            action = (best["date"], best["remark"], best["by"])
        else:
            non_upload_rows = [
                r for r in all_action_rows
                if not _is_upload_remark(r.get("remark", ""))
                and str(r.get("remark", "")).strip() not in ("--", "", "None", "nan")
            ]
            if non_upload_rows:
                best = non_upload_rows[-1]
                print(f"    [DBG] P2 non-upload (of {total}): {best}")
                action = (best["date"], best["remark"], best["by"])
            else:
                best = max(all_action_rows, key=lambda r: str(r.get("date", "")))
                print(f"    [DBG] P3 all-uploads (of {total}), no officer action: {best}")
                action = (best["date"], "No Officer Action Yet", "--")
    else:
        action = ("--", "--", "--")

    try:
        land = await pg.evaluate(_JS_EXTRACT_LAND_DETAILS)
    except Exception as e:
        print(f"    [DBG] land-details eval error: {e}")
        land = None
    land_cols = _flatten_land(land)
    print(f"    [DBG] land details: {land_cols}")
    return action + land_cols


async def fetch_url_view_emutation(context, url: str, sem: asyncio.Semaphore) -> tuple:
    async with sem:
        for attempt in range(3):
            pg = await context.new_page()
            try:
                await pg.route("**/*", _block_all_except_html)
                try:
                    await pg.goto(url, wait_until="domcontentloaded",
                                  timeout=VIEW_GOTO_TIMEOUT)
                except Exception as nav_err:
                    print(f"    [DBG] nav error attempt {attempt+1}: {nav_err}")
                    try: await pg.close()
                    except: pass
                    if attempt < 2:
                        await asyncio.sleep(1.5 * (attempt + 1))
                        continue
                    return ("--",) * 10

                result = await scrape_view_tab_emutation(pg)

                if all(str(x).strip() in ("", "--") for x in result[:3]) and attempt < 2:
                    print(f"    [DBG] blank result attempt {attempt+1}, retrying...")
                    try: await pg.close()
                    except: pass
                    await asyncio.sleep(2.0)
                    continue

                return result
            except Exception as ex:
                print(f"    [DBG] fetch_url_view attempt {attempt+1} exception: {ex}")
                try: await pg.close()
                except: pass
                if attempt < 2:
                    await asyncio.sleep(1.5)
                    continue
                return ("--",) * 10
            finally:
                try: await pg.close()
                except: pass
        return ("--",) * 10


async def fetch_postback_view_emutation(context, detail_page, tr_idx: int) -> tuple:
    try:
        row = detail_page.locator("table tr").nth(tr_idx)
        btn = row.locator("a,input[type=button],button").filter(
            has_text=re.compile(
                r"View|\u0926\u0947\u0916\u0947\u0902|\u0926\u0947\u0916\u0947", re.I)
        ).first
        if await btn.count() == 0:
            btn = row.locator("a,input[type=button],button").first
        if await btn.count() == 0:
            return ("--",) * 10
        async with context.expect_page(timeout=VIEW_GOTO_TIMEOUT) as info:
            await btn.click(delay=50)
        popup = await info.value
        try:
            await popup.route("**/*", _block_all_except_html)
            return await scrape_view_tab_emutation(popup)
        finally:
            try: await popup.close()
            except: pass
    except Exception:
        return ("--",) * 10


# ═══════════════════════════════════════════════════════════════════════════════
# VIEW PAGE SCRAPER — Parimarjan/Rectification style
# ═══════════════════════════════════════════════════════════════════════════════
async def extract_action_parimarjan(pg) -> tuple:
    for _ in range(VIEW_POLL_MAX):
        try:
            r = await pg.evaluate(_JS_TRY_EXTRACT_ACTION)
            if r is not None:
                return r["date"], r["remark"], r["by"]
        except: pass
        await asyncio.sleep(VIEW_POLL_MS / 1000)
    return "--", "--", "--"


async def fetch_url_view_parimarjan(context, url: str, sem: asyncio.Semaphore) -> tuple:
    async with sem:
        pg = await context.new_page()
        try:
            await pg.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,eot,css}", block_assets)
            await pg.goto(url, wait_until="commit", timeout=VIEW_GOTO_TIMEOUT)
            return await extract_action_parimarjan(pg)
        except:
            return "--", "--", "--"
        finally:
            try: await pg.close()
            except: pass


async def fetch_postback_view_parimarjan(context, detail_page, tr_idx: int) -> tuple:
    try:
        row = detail_page.locator("table tr").nth(tr_idx)
        btn = row.locator("a,input[type=button],button").filter(
            has_text=re.compile(
                r"View|\u0926\u0947\u0916\u0947\u0902|\u0926\u0947\u0916\u0947", re.I)
        ).first
        if await btn.count() == 0:
            btn = row.locator("a,input[type=button],button").first
        if await btn.count() == 0:
            return "--", "--", "--"
        async with context.expect_page(timeout=15000) as info:
            await btn.click(delay=50)
        popup = await info.value
        try:
            return await extract_action_parimarjan(popup)
        finally:
            try: await popup.close()
            except: pass
    except:
        return "--", "--", "--"


# ═══════════════════════════════════════════════════════════════════════════════
# UNIFIED DETAIL PAGE SCRAPER
# ═══════════════════════════════════════════════════════════════════════════════
async def scrape_detail_page(context, detail_page, keep, cfg: dict):
    from urllib.parse import urljoin

    scraper_type = cfg["scraper_type"]
    skip_cols    = cfg["skip_cols"]

    if scraper_type == "emutation":
        JS_ROWS      = _JS_EXTRACT_ROWS_EMUTATION
        fetch_url    = fetch_url_view_emutation
        fetch_post   = fetch_postback_view_emutation
        sem          = asyncio.Semaphore(VIEW_CONCURRENCY)
    elif scraper_type == "rectification":
        JS_ROWS      = _JS_EXTRACT_ROWS_RECTIFICATION
        fetch_url    = fetch_url_view_parimarjan
        fetch_post   = fetch_postback_view_parimarjan
        sem          = asyncio.Semaphore(cfg.get("view_concurrency", 8))
    else:
        JS_ROWS      = _JS_EXTRACT_ROWS_PARIMARJAN
        fetch_url    = fetch_url_view_parimarjan
        fetch_post   = fetch_postback_view_parimarjan
        sem          = asyncio.Semaphore(cfg.get("view_concurrency", 8))

    all_rows, headers, page_num = [], [], 1

    try:
        await detail_page.route(
            "**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,eot}", block_assets)
    except: pass

    while True:
        try:
            await detail_page.wait_for_selector("table tr td", timeout=15000)
            extracted = await detail_page.evaluate(JS_ROWS)
        except Exception as e:
            print(f"    Extract error p{page_num}: {e}"); break

        rows_data = extracted.get("rows", [])
        raw_hdrs  = extracted.get("headers", [])

        if not rows_data:
            print(f"    No rows on page {page_num}."); break

        if not headers and raw_hdrs:
            headers = [h for i, h in enumerate(raw_hdrs) if i not in skip_cols]
            headers += ["Last Action Date", "Last Action Remark", "Last Action Taken By"]
            if scraper_type == "emutation":
                headers += ["Halka", "Mauja", "Thana", "Khata No", "Plot No",
                            "Area1(in Acre)", "Area2(in Decimal)"]

        print(f"    Page {page_num}: {len(rows_data)} rows — opening views...")

        url_jobs, post_idx = [], []
        base = detail_page.url

        for i, r in enumerate(rows_data):
            u = r["url"]
            if u:
                if not u.startswith("http"): u = urljoin(base, u)
                url_jobs.append((i, fetch_url(context, u, sem)))
            else:
                post_idx.append(i)

        if url_jobs:
            idxs    = [j[0] for j in url_jobs]
            results = await asyncio.gather(*[j[1] for j in url_jobs],
                                           return_exceptions=True)
            for idx, res in zip(idxs, results):
                rows_data[idx]["action"] = res if not isinstance(res, Exception) \
                                           else ("--", "--", "--")

        for idx in post_idx:
            if len(context.pages) > MAX_OPEN_TABS:
                await close_strays(context, keep)
            rows_data[idx]["action"] = await fetch_post(
                context, detail_page, rows_data[idx]["trIdx"])

        for r in rows_data:
            row = [v for i, v in enumerate(r["cells"]) if i not in skip_cols]
            if scraper_type == "emutation" and not any("/" in str(c) for c in row):
                continue
            act = r.get("action", ("--", "--", "--"))
            row.extend(act)
            all_rows.append(row)

        # ── Detail-page pagination with ellipsis support ─────────────────────────
        next_num = page_num + 1

        async def _find_next_link(pg, num):
            lk = pg.locator(f"a:text-is('{num}')")
            if await lk.count() == 0:
                lk = pg.locator(f"xpath=//a[normalize-space(text())='{num}']")
            return lk if await lk.count() > 0 else None

        next_link = await _find_next_link(detail_page, next_num)

        if next_link is None:
            print(f"    Page {next_num} link not found; checking for ellipsis...")
            try:
                pre_html = await detail_page.evaluate("document.body.innerHTML.length")
            except:
                pre_html = 0

            ellipsis_ok = await _click_ellipsis(detail_page)

            if ellipsis_ok:
                print(f"    Ellipsis triggered — waiting for pager to refresh...")
                await _wait_for_dom_change(detail_page, pre_html)

                # ASP.NET GridView: after ellipsis click, the first page of the next
                # group (next_num) is already loaded and rendered as a <span> (active
                # page indicator), NOT as a clickable <a>.  In that case, the data is
                # already showing — scrape it directly without clicking.
                is_active_span = False
                for _span_poll in range(20):
                    try:
                        is_active_span = await detail_page.evaluate(
                            _JS_ACTIVE_PAGER_PAGE, next_num)
                    except Exception:
                        is_active_span = False
                    if is_active_span:
                        break
                    await asyncio.sleep(0.1)

                if is_active_span:
                    print(f"    Page {next_num} is now active span — data already loaded.")
                    # Data already showing; go back to top of loop to scrape it.
                    # Set page_num to next_num so the loop counter is correct.
                    page_num = next_num
                    next_link = "ALREADY_LOADED"
                else:
                    next_link = await _find_next_link(detail_page, next_num)

            if next_link is None:
                print(f"    Done — no page {next_num} (even after ellipsis check).")
                break

            # Active-span case: data is already loaded, continue to scrape from top.
            # Do NOT increment page_num here — it was already set to next_num above.
            if next_link == "ALREADY_LOADED":
                continue

        try:    pre = await detail_page.locator("table tr").nth(1).inner_text()
        except: pre = ""
        print(f"    -> Page {next_num}...")
        await next_link.first.click()

        for _ in range(75):
            await asyncio.sleep(0.2)
            try:
                if (await detail_page.locator("table tr").nth(1).inner_text()
                        ).strip() != pre.strip():
                    break
            except: break

        if scraper_type == "rectification":
            for _ in range(25):
                await asyncio.sleep(0.2)
                try:
                    first_td = await detail_page.locator("table tr td").nth(1).inner_text()
                    if _VALID_APPID.match(first_td.strip()):
                        break
                except: break

        await asyncio.sleep(0.2)
        page_num += 1

    return all_rows, headers


# ═══════════════════════════════════════════════════════════════════════════════
# BDF BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def _build_bdf(df_list, cfg: dict, for_pdf: bool = False):
    target_report = cfg["report"]
    target_bucket = cfg["bucket"]
    bucket_re     = r'(?<!\d)' + target_bucket + r'(?!\d)'

    dfs = [d for d in df_list
           if d["Report Category"].iloc[0] == target_report
           and d["Case Type"].astype(str).str.contains(bucket_re, regex=True).any()]
    if not dfs:
        return None

    cat = pd.concat(dfs, ignore_index=True)
    bdf = cat[cat["Case Type"].astype(str).str.contains(bucket_re, regex=True)].copy()
    if bdf.empty:
        return None

    if for_pdf and cfg["remark_filter"] and "Last Action Remark" in bdf.columns:
        bdf = bdf[~_remark_should_exclude(bdf["Last Action Remark"])].copy()
        if bdf.empty:
            return None

    return bdf


# ═══════════════════════════════════════════════════════════════════════════════
# PDF GENERATION
# ═══════════════════════════════════════════════════════════════════════════════
PDF_COL_WIDTHS_CM = {
    "Sr. No.":               1.5,
    "Token No":              4.0,
    "Date of Application":   3.2,
    "Last Action Date":      3.2,
    "Last Action Remark":    7.0,
    "Last Action Taken By":  4.5,
    "Halka":                 3.0,
    "Case Type":             3.0,
}
PDF_DEFAULT_COL_CM = 3.0

def para(txt, st):
    return Paragraph(str(txt) if txt else "", st)

def generate_pdf(master: dict, out_dir: str, cfg: dict) -> dict:
    os.makedirs(out_dir, exist_ok=True)
    paths = {}

    ts    = ParagraphStyle("t",   fontName=_LATB, fontSize=14,
                           textColor=colors.HexColor("#1F4E79"), spaceAfter=4)
    ss    = ParagraphStyle("s",   fontName=_LATB, fontSize=11,
                           textColor=colors.HexColor("#4F81BD"), spaceAfter=4)
    ms    = ParagraphStyle("m",   fontName=_LAT,  fontSize=9,
                           textColor=colors.grey,               spaceAfter=8)
    th_st = ParagraphStyle("th",  fontName=_LATB, fontSize=8,
                           textColor=colors.white, leading=12)
    td_st = ParagraphStyle("td",  fontName=_LAT,  fontSize=8,
                           textColor=colors.black, leading=11)
    td_alt = ParagraphStyle("tda", fontName=_LAT, fontSize=8,
                            textColor=colors.black, leading=11)

    target_report = cfg["report"]
    pending_label = cfg["pending_label"]
    file_tag      = cfg["file_tag"]
    remark_filter = cfg["remark_filter"]

    for karm, df_list in master.items():
        bdf = _build_bdf(df_list, cfg, for_pdf=True)
        if bdf is None:
            continue

        sk    = re.sub(r'[^\w\s-]', '', karm).strip().replace(' ', '_')
        fname = f"{sk}_{file_tag}.pdf"
        fpath = os.path.join(out_dir, fname)

        display_cols = [c for c in bdf.columns
                        if c not in ("Report Category", "_bucket", "_r")]

        doc = SimpleDocTemplate(fpath, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm,  bottomMargin=1.5*cm)

        tbl_rows = [[para(c, th_st) for c in display_cols]]
        for ri, (_, row) in enumerate(bdf[display_cols].iterrows()):
            st = td_alt if ri % 2 else td_st
            tbl_rows.append([para(v, st) for v in row])

        usable_width = 27.0
        col_w_list = [PDF_COL_WIDTHS_CM.get(c, PDF_DEFAULT_COL_CM) * cm
                      for c in display_cols]
        total_w = sum(col_w_list)
        if total_w > 0:
            scale = (usable_width * cm) / total_w
            col_w_list = [w * scale for w in col_w_list]

        tbl = Table(tbl_rows, colWidths=col_w_list, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0,0),(-1,0),  colors.HexColor("#4F81BD")),
            ("TEXTCOLOR",      (0,0),(-1,0),  colors.white),
            ("FONTNAME",       (0,0),(-1,0),  _LATB),
            ("FONTSIZE",       (0,0),(-1,-1), 8),
            ("ALIGN",          (0,0),(-1,-1), "CENTER"),
            ("VALIGN",         (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",     (0,0),(-1,-1), 4),
            ("BOTTOMPADDING",  (0,0),(-1,-1), 4),
            ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#F0F4FA")]),
            ("GRID",           (0,0),(-1,-1), 0.4, colors.grey),
            ("FONTNAME",       (0,1),(-1,-1), _LAT),
        ]))

        meta_line = (
            f"Cases Pending > {pending_label} Days  |  Total: {len(bdf)}  |  "
            f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        )
        if remark_filter:
            excl_note = (
                "  |  Excluded: "
                + ", ".join(f'"{x}"' for x in sorted(EXCLUDED_EXACT))
                + " | NaN/empty | Starts-with: "
                + ", ".join(f'"{p}..."' for p in EXCLUDED_PREFIXES)
            )
            meta_line += excl_note

        story = [
            para(f"Karmchari: {karm}", ts),
            para(target_report, ss),
            para(meta_line, ms),
            HRFlowable(width="100%", thickness=1,
                       color=colors.HexColor("#4F81BD"), spaceAfter=8),
            tbl,
            Spacer(1, 0.5*cm),
            para(f"End of report — {len(bdf)} cases.", ms),
        ]
        try:
            doc.build(story)
            paths[karm] = fpath
            print(f"  PDF: {fname} ({len(bdf)} rows)")
        except Exception as e:
            print(f"  PDF error {karm}: {e}")

    return paths


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITERS  (single combined workbook)
# ═══════════════════════════════════════════════════════════════════════════════

def _write_one_report_block(ws, current_row: int, df_list, cfg: dict) -> int:
    """
    Write one report/bucket block into ws starting at current_row.
    Returns the next free row after the block.
    """
    HF  = PatternFill("solid", fgColor="4F81BD")
    HN  = Font(bold=True, color="FFFFFF")
    SF  = Font(bold=True, size=11, color="FFFFFF")
    SFL = PatternFill("solid", fgColor="1F4E79")
    BRD = Border(*[Side(style="thin")]*4)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT = PatternFill("solid", fgColor="EBF3FB")

    # ── Title row for this block ─────────────────────────────────────────────
    title_text = f"  {cfg['report'].upper()} — PENDING > {cfg['pending_label']} DAYS  "
    title_cell = ws.cell(current_row, 1, title_text)
    title_cell.font      = SF
    title_cell.fill      = SFL
    title_cell.alignment = CTR
    # We'll merge across columns after we know how wide the data is
    title_row = current_row
    current_row += 1

    bdf = _build_bdf(df_list, cfg, for_pdf=False)

    if bdf is None or bdf.empty:
        no_data = ws.cell(current_row, 1, "No pending cases found.")
        no_data.font = Font(italic=True, color="808080")
        ws.cell(current_row, 2, 0).font = Font(bold=True)
        current_row += 2   # blank separator
        return current_row

    df = bdf.drop(columns=["Report Category"], errors="ignore")
    nc = len(df.columns)

    # Merge title across all data columns
    if nc > 1:
        ws.merge_cells(start_row=title_row, start_column=1,
                       end_row=title_row, end_column=nc)

    # ── Header row ───────────────────────────────────────────────────────────
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(current_row, ci, col)
        cell.fill, cell.font, cell.alignment, cell.border = HF, HN, CTR, BRD
    header_row = current_row
    current_row += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(current_row, ci, val)
            cell.fill      = ALT if ri % 2 == 0 else PatternFill(fill_type=None)
            cell.font      = Font(color="000000")
            cell.border    = BRD
            cell.alignment = CTR
        current_row += 1

    # ── Column widths (only grow, never shrink) ──────────────────────────────
    for ci, col in enumerate(df.columns, 1):
        mx = df[col].astype(str).str.len().max()
        w  = max(mx if pd.notnull(mx) else 0, len(str(col))) + 2
        col_letter = get_column_letter(ci)
        existing = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = min(max(existing, w), 50)

    current_row += 1   # blank separator row between blocks
    return current_row


def write_all_reports_for_karm(ws, all_masters: list, karm: str):
    """
    Write all 12 report blocks one after another into ws for this karmchari.
    all_masters: list of (master_dict, cfg) tuples, one per REPORT_CONFIGS entry.
    """
    current_row = 1
    for master, cfg in all_masters:
        df_list = master.get(karm, [])
        current_row = _write_one_report_block(ws, current_row, df_list, cfg)


def build_combined_summary(wb, all_masters: list, pdf_path_map: dict = None):
    """
    Build a single Summary sheet covering all 12 reports.
    Columns: Karmchari | Report (tag) | Count   ... one group per report config.
    Layout: report configs as column-groups, karmcharis as rows.

    pdf_path_map: optional dict  {file_tag: {karm: abs_pdf_path}}
    When provided, non-zero count cells become hyperlinks that open the PDF.
    """
    if pdf_path_map is None:
        pdf_path_map = {}

    ws  = wb.create_sheet("Summary", 0)
    TF  = Font(bold=True, size=13, color="FFFFFF")
    TL  = PatternFill("solid", fgColor="1F4E79")
    HF  = Font(bold=True, color="FFFFFF")
    HL  = PatternFill("solid", fgColor="4F81BD")
    NL  = PatternFill("solid", fgColor="D6E4F0")
    NF  = Font(bold=True)
    XL  = PatternFill("solid", fgColor="FCE4D6")
    XF  = Font(bold=True)
    BRD = Border(*[Side(style="thin")]*4)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

    num_reports = len(all_masters)   # 12

    # ── Row 1: Main title ────────────────────────────────────────────────────
    title = ws.cell(1, 1, "  COMBINED SUMMARY — ALL REPORTS  ")
    title.font, title.fill, title.alignment = TF, TL, CTR
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=1 + num_reports)

    # ── Row 2: Column headers ────────────────────────────────────────────────
    hdr0 = ws.cell(2, 1, "Karmchari Name")
    hdr0.font, hdr0.fill, hdr0.alignment, hdr0.border = HF, HL, CTR, BRD

    for col_i, (_, cfg) in enumerate(all_masters, 2):
        label = f"{cfg['file_tag'].replace('_', ' ')}\n(>{cfg['pending_label']}d)"
        hc = ws.cell(2, col_i, label)
        hc.font, hc.fill, hc.alignment, hc.border = HF, HL, CTR, BRD

    total_hdr = ws.cell(2, 2 + num_reports, "Grand Total")
    total_hdr.font, total_hdr.fill, total_hdr.alignment, total_hdr.border = HF, HL, CTR, BRD

    ws.row_dimensions[2].height = 40

    # ── Rows 3+: One row per karmchari ──────────────────────────────────────
    grand_totals = [0] * num_reports
    for row_i, karm in enumerate(ALL_KARMCHARIS, 3):
        # Karmchari name cell — hyperlink to their sheet
        sn_safe = karm[:31].replace("'", "''")
        nc = ws.cell(row_i, 1, karm)
        nc.font, nc.fill, nc.alignment, nc.border = NF, NL, CTR, BRD
        nc.hyperlink = f"#'{sn_safe}'!A1"
        nc.font = Font(bold=True, underline="single", color="1F4E79")

        row_total = 0
        for col_i, (master, cfg) in enumerate(all_masters, 2):
            df_list = master.get(karm, [])
            bdf = _build_bdf(df_list, cfg, for_pdf=True)
            cnt = len(bdf) if bdf is not None else 0

            BG, FG = cfg["bucket_color"]
            cnt_cell = ws.cell(row_i, col_i, cnt if cnt else 0)
            cnt_cell.alignment, cnt_cell.border = CTR, BRD
            if cnt > 0:
                cnt_cell.fill = PatternFill("solid", fgColor=BG)
                # Add PDF hyperlink if we have a path for this karm+report
                pdf_for_tag  = pdf_path_map.get(cfg["file_tag"], {})
                pdf_abs_path = pdf_for_tag.get(karm)
                if pdf_abs_path and os.path.exists(pdf_abs_path):
                    # Convert to file:// URI (works on Windows and Linux)
                    uri = "file:///" + pdf_abs_path.replace("\\", "/").lstrip("/")
                    cnt_cell.hyperlink = uri
                    cnt_cell.font = Font(color=FG, bold=True, underline="single")
                else:
                    cnt_cell.font = Font(color=FG, bold=True)
            else:
                cnt_cell.font = Font(color="AAAAAA")
                cnt_cell.fill = PatternFill(fill_type=None)

            grand_totals[col_i - 2] += cnt
            row_total += cnt

        # Row total
        rt_cell = ws.cell(row_i, 2 + num_reports, row_total)
        rt_cell.font      = XF
        rt_cell.alignment = CTR
        rt_cell.border    = BRD
        if row_total > 0:
            rt_cell.fill = PatternFill("solid", fgColor="FCE4D6")

    # ── TOTAL row ────────────────────────────────────────────────────────────
    total_row = 3 + len(ALL_KARMCHARIS)
    ws.cell(total_row, 1, "TOTAL").font = XF
    ws.cell(total_row, 1).fill         = XL
    ws.cell(total_row, 1).alignment    = CTR
    ws.cell(total_row, 1).border       = BRD

    grand_grand = 0
    for col_i, cnt in enumerate(grand_totals, 2):
        tc = ws.cell(total_row, col_i, cnt)
        tc.font, tc.fill, tc.alignment, tc.border = XF, XL, CTR, BRD
        grand_grand += cnt

    gg = ws.cell(total_row, 2 + num_reports, grand_grand)
    gg.font, gg.fill, gg.alignment, gg.border = XF, XL, CTR, BRD

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    for col_i in range(2, 2 + num_reports + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 18


# ═══════════════════════════════════════════════════════════════════════════════
# COMBINED EXCEL + PDF OUTPUT WRITER
# ═══════════════════════════════════════════════════════════════════════════════
def write_combined_outputs(all_masters: list):
    """
    all_masters: list of (master_dict, cfg) tuples — one per REPORT_CONFIGS entry.

    Produces:
      • One Excel file on the Desktop with:
          – One sheet per karmchari, containing all 12 report blocks stacked
          – A combined Summary sheet (all 12 reports × all karmcharis)
      • One PDF folder per report/bucket (unchanged behaviour).
    """
    from openpyxl import Workbook

    desk = os.path.expanduser("~/Desktop")
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = os.path.join(desk, f"Karmchari_AllReports_{ts}.xlsx")

    # ── Generate PDFs (unchanged per-report logic) ───────────────────────────
    # pdf_path_map: {file_tag: {karm: abs_path}}  — used later for hyperlinks
    all_pdf_paths = {}   # karm  -> {file_tag -> path}  (legacy, not used below)
    pdf_path_map  = {}   # file_tag -> {karm  -> path}  — for Summary hyperlinks
    for master, cfg in all_masters:
        if not master:
            continue
        pdf_dir   = os.path.join(desk, f"Karmchari_PDFs_{cfg['file_tag']}")
        print(f"\nGenerating PDFs for {cfg['file_tag']} -> {pdf_dir}")
        if cfg["remark_filter"]:
            print(f"  Excluded exact : {sorted(EXCLUDED_EXACT)}")
            print(f"  Excluded prefix: {list(EXCLUDED_PREFIXES)}")
        pdf_paths = generate_pdf(master, pdf_dir, cfg)
        print(f"  {len(pdf_paths)} PDFs done.")
        for karm, path in pdf_paths.items():
            all_pdf_paths.setdefault(karm, {})[cfg["file_tag"]] = path
        pdf_path_map[cfg["file_tag"]] = dict(pdf_paths)

    # ── Build single Excel workbook ──────────────────────────────────────────
    print(f"\nGenerating combined Excel -> {xlsx}")
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # One sheet per karmchari (all 12 report blocks stacked)
    for karm in ALL_KARMCHARIS:
        sn = karm[:31]
        ws = wb.create_sheet(sn)
        write_all_reports_for_karm(ws, all_masters, karm)
        print(f"  Sheet written: {sn}")

    # Combined summary sheet (inserted at position 0)
    build_combined_summary(wb, all_masters, pdf_path_map=pdf_path_map)

    # Move Summary to first position
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    wb.save(xlsx)
    print(f"  Excel saved -> {xlsx}")


def write_live_excel(all_masters: list, xlsx: str | None = None):
    """
    Lightweight Excel-only writer that can be called repeatedly while scraping.
    It does NOT regenerate PDFs; it only mirrors the current all_masters state
    into an Excel workbook at a fixed path (overwriting if it already exists).
    """
    from openpyxl import Workbook

    desk = os.path.expanduser("~/Desktop")
    if xlsx is None:
        xlsx = os.path.join(desk, "Karmchari_AllReports_LIVE.xlsx")

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for karm in ALL_KARMCHARIS:
        sn = karm[:31]
        ws = wb.create_sheet(sn)
        write_all_reports_for_karm(ws, all_masters, karm)

    build_combined_summary(wb, all_masters)

    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    # Write to a temp file first, then atomically replace the target so
    # partially-written files are never left behind. If the main file is
    # open/locked in Excel, fall back to *_latest.xlsx.
    tmp = xlsx + ".tmp"
    wb.save(tmp)
    try:
        os.replace(tmp, xlsx)
        print(f"[CHKPT] Live Excel saved -> {xlsx}")
    except PermissionError:
        alt = xlsx.replace(".xlsx", "_latest.xlsx")
        try:
            os.replace(tmp, alt)
            print(f"[CHKPT] Main file locked; saved -> {alt}")
        except Exception:
            if os.path.exists(tmp):
                os.remove(tmp)
            raise
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


# ═══════════════════════════════════════════════════════════════════════════════
# SINGLE REPORT RUNNER
# ═══════════════════════════════════════════════════════════════════════════════
async def run_single_report(context, page, dash, cfg: dict,
                            all_masters: list | None = None,
                            live_xlsx: str | None = None) -> dict:
    SIDEBAR       = "xpath=/html/body/form/div[3]/div/aside/div/nav/ul/li[12]/a"
    target_bucket = cfg["bucket"]
    target_report = cfg["report"]
    aging_extra   = cfg["aging_extra"]
    col_fallback  = cfg["col_fallback"]
    file_tag      = cfg["file_tag"]
    pending_label = cfg["pending_label"]
    master        = {}

    print(f"\n{'='*60}")
    print(f"--- Processing: {target_report}  (>{pending_label} Days) ---")

    try:
        dash = await live_dashboard(context, dash)
        await dismiss_dialogs(dash)

        if not await goto_reports(dash, SIDEBAR):
            print("Sidebar failed — checking session...")
            dash = await ensure_session(context, page, dash)
            if not await goto_reports(dash, SIDEBAR):
                print("Sidebar still failing, skipping this report.")
                return master

        report_dd = dash.locator(
            "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div/div[1]/select")
        await report_dd.wait_for(state="visible", timeout=30000)
        for _ in range(15):
            if not await report_dd.is_disabled() and \
               await report_dd.locator("option").count() > 1:
                break
            await asyncio.sleep(0.5)

        all_opts    = await report_dd.locator("option").all_inner_texts()
        search_term = target_report.lower().replace(" ", "")
        target_opt  = next(
            (o for o in all_opts if search_term in o.lower().replace(" ", "")), None)
        if target_opt:
            await report_dd.select_option(label=target_opt.strip())
        else:
            await report_dd.select_option(label=target_report)
        await asyncio.sleep(2)

        await dash.select_option(
            "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div/div[2]/select",
            label="Patna")
        await asyncio.sleep(2)
        await dash.select_option(
            "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div/div[3]/select",
            label="Masaurhi")
        await asyncio.sleep(2)

        PROC   = "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div/div[4]/input[1]"
        GENBT  = "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div[2]/div[4]/input"
        TBLX   = "xpath=/html/body/form/div[3]/div/section/div/div/div/div[3]/div[1]/table"
        FROM_X = "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div[2]/div[2]/input[1]"
        TO_X   = "xpath=/html/body/form/div[3]/div/section/div/div/div/div[1]/div[2]/div[2]/div[3]/input[1]"

        found = False
        for attempt in range(10):
            await dismiss_dialogs(dash)
            await dash.click(PROC)
            try:
                await dash.wait_for_selector(FROM_X, timeout=15000)
                today    = datetime.now().strftime("%d-%m-%Y")
                from_box = dash.locator(FROM_X)
                to_box   = dash.locator(TO_X)
                await from_box.evaluate("el => el.removeAttribute('readonly')")
                await to_box.evaluate("el => el.removeAttribute('readonly')")
                await from_box.fill("01-04-2025")
                await to_box.fill(today)
                print(f"  Generating summary (attempt {attempt+1})...")
                await dash.click(GENBT)
                await dash.wait_for_selector(TBLX, timeout=25000)
                for _ in range(10):
                    if await dash.locator(f"{TBLX}/tbody/tr[position()>3]").count() > 0:
                        found = True; break
                    await asyncio.sleep(1)
                if found: break
            except:
                print(f"  Attempt {attempt+1} failed.")
                await asyncio.sleep(3)

        if not found:
            print("Summary table not found, skipping.")
            return master

        sum_hdrs = [h.strip() for h in await dash.locator(
            f"{TBLX}/tbody/tr").nth(2).locator("th,td").all_inner_texts()]

        if aging_extra:
            aging_cols = [
                i for i, h in enumerate(sum_hdrs)
                if (">" in h or "अधिक" in h)
                and re.search(r'(?<!\d)' + target_bucket + r'(?!\d)', h)
                and "total" not in h.lower() and "कुल" not in h
            ]
        else:
            aging_cols = [
                i for i, h in enumerate(sum_hdrs)
                if re.search(r'(?<!\d)' + target_bucket + r'(?!\d)', h)
                and "total" not in h.lower() and "कुल" not in h
            ]

        if not aging_cols:
            aging_cols = [col_fallback]
            print(f"  Auto-detect failed; using fallback column index {col_fallback}.")
        else:
            print(f"  {pending_label}-day column(s) found at indices: {aging_cols}")

        rows      = dash.locator(f"{TBLX}/tbody/tr[position()>3]")
        row_count = await rows.count()
        print(f"  Summary rows: {row_count}")

        for i in range(row_count):
            halka     = (await rows.nth(i).locator("td").nth(2).inner_text()).strip()
            karmchari = HALKA_MAPPING.get(halka, HALKA_MAPPING.get(halka.upper(), "Others"))

            for ci in aging_cols:
                hdr = sum_hdrs[ci].lower() if ci < len(sum_hdrs) else ""
                if "total" in hdr or "कुल" in hdr: continue

                cell  = rows.nth(i).locator("td").nth(ci)
                text  = await cell.inner_text()
                cases = int("".join(filter(str.isdigit, text))) if any(c.isdigit() for c in text) else 0
                if cases == 0: continue

                ctype = sum_hdrs[ci]
                print(f"  {halka} -> {karmchari} | {cases} cases | {ctype}")

                # Check session before opening detail page
                dash = await ensure_session(context, page, dash)

                detail = None
                for da in range(10):
                    try:
                        async with context.expect_page(timeout=120000) as di:
                            lnk = cell.locator("a")
                            if await lnk.count() > 0: await lnk.click()
                            else: await cell.click()
                        detail = await di.value
                        await detail.wait_for_load_state("domcontentloaded", timeout=60000)
                        # Check if detail page itself redirected to logout
                        if await is_session_expired(detail):
                            await detail.close()
                            detail = None
                            dash = await re_login(context, page)
                            break
                        await detail.wait_for_selector("table tr td", timeout=30000)
                        break
                    except Exception as ep:
                        print(f"    Detail attempt {da+1}: {ep}")
                        if detail:
                            try: await detail.close()
                            except: pass
                            detail = None
                        await asyncio.sleep(1)

                if not detail:
                    print("    Could not open detail page."); continue

                await dismiss_dialogs(detail)
                keep = [page, dash, detail]

                try:
                    all_rows, hdrs = await scrape_detail_page(
                        context, detail, keep, cfg)
                    if all_rows and hdrs:
                        df = pd.DataFrame(all_rows)
                        while len(df.columns) < len(hdrs):
                            df[f"x{len(df.columns)}"] = ""
                        df = df.iloc[:, :len(hdrs)]
                        df.columns = hdrs
                        df = df.drop(
                            columns=[c for c in df.columns if any(
                                x in str(c).lower()
                                for x in ["total", "कुल", "sum"])],
                            errors="ignore")
                        df["Halka"]           = halka
                        df["Report Category"] = target_report
                        df["Case Type"]       = ctype
                        master.setdefault(karmchari, []).append(df)
                        print(f"    {len(all_rows)} rows -> {karmchari}")
                        if all_masters is not None and live_xlsx:
                            try:
                                write_live_excel(all_masters, live_xlsx)
                            except Exception as e:
                                print(f"[CHKPT] Live Excel write failed: {e}")
                    else:
                        print("    No rows scraped.")
                except Exception as e:
                    print(f"    Scrape error: {e}")
                finally:
                    try: await detail.close()
                    except: pass
                    gc.collect()

    except Exception as e:
        print(f"Report block error: {e}")

    return master





# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
async def run():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page    = await context.new_page()

        for attempt in range(3):
            try:
                print(f"--- Navigating to Portal (Attempt {attempt+1}) ---")
                await page.goto(
                    "https://biharbhumi.bihar.gov.in/BiharEmutationNew/",
                    wait_until="load", timeout=90000)
                break
            except Exception as e:
                if attempt == 2: raise e
                await asyncio.sleep(5)

        await page.select_option("#cbo_District_Code", label="State")
        await page.wait_for_load_state("domcontentloaded")
        await page.select_option("#cbo_Circle_Code1", label="Bihar")
        await page.fill("#txt_UserName", "Bihar")
        await page.fill("#txt_Password1", "nic123")
        n1 = await page.locator("#txtNum1").inner_text()
        n2 = await page.locator("#txtNum2").inner_text()
        await page.fill("#txtResult", str(int(n1.strip()) + int(n2.strip())))
        await dismiss_dialogs(page)

        async with context.expect_page() as npi:
            await page.click("#btn_Login")
        dash = await npi.value
        await dash.wait_for_load_state("load")
        await dismiss_dialogs(dash)
        print("Login OK")

        desk = os.path.expanduser("~/Desktop")
        live_xlsx = os.path.join(desk, "Karmchari_AllReports_LIVE.xlsx")

        all_masters = []
        for cfg in REPORT_CONFIGS:
            dash = await live_dashboard(context, dash)
            await dismiss_dialogs(dash)

            # Ensure session is alive before starting each report
            dash = await ensure_session(context, page, dash)

            master = {}
            all_masters.append((master, cfg))
            await run_single_report(context, page, dash, cfg,
                                    all_masters=all_masters,
                                    live_xlsx=live_xlsx)
            dash = await live_dashboard(context, dash)

        # Write everything into one combined Excel file
        write_combined_outputs(all_masters)

        print("\nALL REPORTS DONE.")
        await browser.close()


if __name__ == "__main__":
    asyncio.run(run())
